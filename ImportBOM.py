import sys
import json
import os
import pandas as pd
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QDragEnterEvent, QDropEvent, QFont
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
import traceback

def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(__file__)

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(__file__)
    return os.path.join(base, relative_path)

class BOMPage(QtWidgets.QWidget):
    def __init__(self, script_dir):
        super().__init__()
        self.script_dir = script_dir
        self.init_ui()

    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.label = QtWidgets.QLabel("将 Excel 文件拖拽到此处或点击浏览按钮", alignment=QtCore.Qt.AlignCenter)
        self.label.setStyleSheet("border: 2px dashed #aaa; padding: 40px;")
        layout.addWidget(self.label)
        self.browse_btn = QtWidgets.QPushButton("浏览文件")
        layout.addWidget(self.browse_btn)
        self.browse_btn.clicked.connect(self.open_file_dialog)
        self.text_edit = QtWidgets.QTextEdit()
        self.text_edit.setReadOnly(True)
        layout.addWidget(self.text_edit, stretch=1)
        self.setAcceptDrops(True)

    def load_mapping(self):
        json_path = os.path.join(self.script_dir, "BillofMaterial.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                json_list = json.load(f)
        except Exception as e:
            full_tb = traceback.format_exc()
            QMessageBox.critical(self, "错误", f"加载映射文件出错：{e}\n\n{full_tb}")
            json_list = []
        else:
            QMessageBox.information(self, "提示", f"成功加载 {len(json_list)} 条映射文件")
        self.mapping = {
            (item.get("Value", ""), item.get("PCB Decal", "")):
                (item.get("Material Description", ""), item.get("Type", ""))
            for item in json_list
        }

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls(): event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if not urls: return
        path = urls[0].toLocalFile()
        if not path.lower().endswith(('.xlsx', '.xls')):
            self.text_edit.append("仅支持 .xls/.xlsx 文件")
            QMessageBox.warning(self, "警告", "仅支持 .xls/.xlsx 文件")
            return
        self.process_file(path)

    def open_file_dialog(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.process_file(path)

    def process_file(self, path):
        try:
            raw = pd.read_excel(path, header=None, dtype=str)
            target = ["Item", "Quantity", "Reference", "Part", "PCB Footprint"]
            header_idx = next((i for i,row in raw.iterrows()
                               if list(row.iloc[:5].fillna("").str.strip())==target), None)
            if header_idx is None:
                raise ValueError("未找到表头行：Item Quantity Reference Part PCB Footprint")
            second = raw.iloc[header_idx,:5].tolist()
            body = raw.iloc[header_idx+1:,:5].reset_index(drop=True).fillna("").astype(str)
            body.columns = pd.MultiIndex.from_product([["Bill of Material"], second])
            # ===== 删除 Value(Part) 列中以 NC/ 开头的物料 =====
            part_col = ("Bill of Material", "Part")
            body = body[
                ~body[part_col]
                    .fillna("")
                    .str.upper()
                    .str.startswith("NC")
            ]
            # ==============================================
            def lookup(r, idx):
                key=(r[("Bill of Material","Part")].strip(),
                     r[("Bill of Material","PCB Footprint")].strip())
                return self.mapping.get(key, ("", ""))[idx]
            body[("Bill of Material","Material Description")] = body.apply(lambda r: lookup(r,0),axis=1)
            body[("Bill of Material","Type")] = body.apply(lambda r: lookup(r,1),axis=1)
            out=os.path.splitext(path)[0] + "_output.xlsx"
            body.to_excel(out, engine='openpyxl')
            wb = load_workbook(out)
            ws = wb.active
            ws.delete_rows(3,2); ws.delete_cols(1,1)
            data=[list(r) for r in ws.iter_rows(values_only=True)]
            for r in data:
                r[1],r[6]=r[6],r[1]; r[2],r[5]=r[5],r[2]; r.append(r.pop(5))
            ws.delete_rows(1,ws.max_row)
            for i,row in enumerate(data,start=1):
                for j,v in enumerate(row,start=1): ws.cell(i,j,v)
            ws.delete_rows(1,2); ws.insert_rows(1); ws.insert_rows(1)
            for c in range(1,8): ws.cell(1,c,"Bill of Material")
            heads=["No.","Type","Material Description","Value","PCB Decal","Quantity","Reference"]
            for i,h in enumerate(heads,start=1): ws.cell(2,i,h)
            rows=[list(r) for r in ws.iter_rows(values_only=True)]
            idxs=list(range(3,len(rows)))
            sorted_idx=sorted(idxs,key=lambda i:rows[i][1] or 'z')
            ws.delete_rows(4,ws.max_row-3)
            for pos,i in enumerate(sorted_idx,start=4):
                for col,v in enumerate(rows[i],start=1): ws.cell(pos,col,v)
            dims={'A':5,'C':25,'D':14,'E':18,'F':8,'G':25}
            for c,w in dims.items(): ws.column_dimensions[c].width=w
            for r in range(3,ws.max_row+1): ws.row_dimensions[r].height=28
            for r in (1,2): ws.row_dimensions[r].height=30
            img=XLImage(resource_path("Title.png")); img.width,img.height=850,40; ws.add_image(img,"A1")
            for row in ws.iter_rows(1,ws.max_row,1,ws.max_column):
                for cell in row:
                    if cell.value is not None: cell.alignment=Alignment(wrap_text=True)
            for row in ws.iter_rows(1,2,1,ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.font=Font(bold=True)
            start,prev=3,ws.cell(3,2).value
            for r in range(4,ws.max_row+1):
                curr=ws.cell(r,2).value
                if curr!=prev:
                    ws.merge_cells(start_row=start,start_column=2,end_row=r-1,end_column=2)
                    start,prev=r,curr
            wb.save(out)
            QMessageBox.information(self,"完成","Excel 导出成功")
            self.text_edit.append("Excel 导出成功")
        except Exception as e:
            tb=traceback.format_exc()
            QMessageBox.critical(self,"错误",f"处理失败：{e}\n\n{tb}")
            self.text_edit.setPlainText(str(e))

class JSONUpdatePage(QtWidgets.QWidget):
    def __init__(self,script_dir):
        super().__init__();self.script_dir=script_dir
        self.init_ui();self.setAcceptDrops(True)

    def init_ui(self):
        layout=QtWidgets.QVBoxLayout(self)
        self.label=QtWidgets.QLabel("将 JSON/Excel 拖拽或点击浏览更新映射",alignment=QtCore.Qt.AlignCenter)
        self.label.setStyleSheet("border:2px dashed #aaa;padding:40px;")
        layout.addWidget(self.label)
        self.browse_btn=QtWidgets.QPushButton("浏览文件")
        layout.addWidget(self.browse_btn)
        self.browse_btn.clicked.connect(self.open_file_dialog)
        self.status=QtWidgets.QLabel("等待更新...",alignment=QtCore.Qt.AlignCenter)
        font=QFont();font.setBold(True);self.status.setFont(font)
        layout.addWidget(self.status)

    def dragEnterEvent(self,event):
        if event.mimeData().hasUrls():event.acceptProposedAction()
    def dropEvent(self,event):
        urls=event.mimeData().urls();
        if not urls: return
        self.process_update(urls[0].toLocalFile())

    def open_file_dialog(self):
        path,_=QFileDialog.getOpenFileName(self,"选择映射文件","","All Files (*.json *.xlsx *.xls)")
        if path: self.process_update(path)
    def process_update(self, path):
        try:
            ext = os.path.splitext(path)[1].lower()
            script = os.path.join(self.script_dir, 'BillofMaterial.json')

            if ext == '.json':
                data = json.load(open(path, 'r', encoding='utf-8'))
            else:
                # 先读取所有 sheet
                sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=str)
                headers = ['Type', 'Material Description', 'Value', 'PCB Decal']
                records = []

                for raw in sheets.values():
                    # 找到表头所在行号
                    idx = next(
                        (i for i, row in raw.iterrows()
                         if set(headers).issubset([str(c).strip() for c in row.tolist()])),
                        None
                    )
                    if idx is None:
                        continue

                    # 用 raw 构造 df，并将 idx 行设为列名
                    df = raw.copy()
                    df.columns = df.iloc[idx].tolist()
                    df = df.iloc[idx + 1:]

                    # 只保留目标列，删除全空行，填充空字符串
                    df = df[headers].dropna(how='all').fillna('')
                    # 向下填充 Type 列
                    df['Type'] = df['Type'].replace('', pd.NA).ffill()

                    records.extend(df.to_dict('records'))

                if not records:
                    raise ValueError('无有效映射数据')

                data = records

            # 写回 JSON
            with open(script, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.status.setText(f"更新成功，共 {len(data)} 条记录。")
            QMessageBox.information(self, "完成", f"JSON 更新成功：{len(data)} 条")

        except Exception as e:
            tb = traceback.format_exc()
            QMessageBox.critical(self, "错误", f"更新失败：{e}\n\n{tb}")
            self.status.setText(f"更新失败：{e}")

class MainWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__();self.setWindowTitle("BOM处理工具");self.resize(1000,600)
        dir_=get_app_dir();self.sidebar=QtWidgets.QListWidget();self.sidebar.addItems(["BOM 增加描述","更新硬件库"])
        self.sidebar.setFixedWidth(150)
        self.page1=BOMPage(dir_);self.page2=JSONUpdatePage(dir_)
        self.stack=QtWidgets.QStackedWidget();self.stack.addWidget(self.page1);self.stack.addWidget(self.page2)
        lay=QtWidgets.QHBoxLayout(self);lay.addWidget(self.sidebar);lay.addWidget(self.stack,1)
        self.sidebar.currentRowChanged.connect(self.on_change);self.sidebar.setCurrentRow(0)
    def on_change(self, index):
        self.stack.setCurrentIndex(index)
        if index == 0:
            self.page1.load_mapping()

if __name__=='__main__':
    app=QtWidgets.QApplication(sys.argv);app.setFont(QFont("Arial",12))
    w=MainWindow();w.show();sys.exit(app.exec_())
